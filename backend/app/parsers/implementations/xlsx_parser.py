"""XLSX workbook parser using openpyxl."""

from __future__ import annotations

from pathlib import Path
from time import perf_counter
from typing import Any

from openpyxl import load_workbook

from app.parsers.base import BaseParser, ParseResult, ParserInfo
from app.parsers.registry import ParserRegistry

SCHEMA_VERSION = "1.0"


def _cell_to_string(value: Any) -> str:
    """Convert a workbook cell value to a display string."""
    if value is None:
        return ""
    return str(value)


class XlsxParser(BaseParser):
    """Parse XLSX workbooks into unified document blocks."""

    def info(self) -> ParserInfo:
        return ParserInfo(
            name="xlsx_parser",
            supported_types=["xls", "xlsx"],
            required_gpu=False,
            required_models=[],
            version="1.0.0",
        )

    def parse(
        self,
        file_path: str,
        options: dict[str, Any] | None = None,
    ) -> ParseResult:
        started = perf_counter()
        path = Path(file_path)
        if not path.is_file():
            raise FileNotFoundError(file_path)

        workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
        blocks: list[dict[str, Any]] = []
        try:
            for sheet_index, sheet in enumerate(workbook.worksheets, start=1):
                blocks.append(
                    {
                        "type": "heading",
                        "level": 2,
                        "text": f"Sheet {sheet.title}",
                        "page": sheet_index,
                    }
                )
                rows = []
                for row in sheet.iter_rows(values_only=True):
                    rows.append([_cell_to_string(cell) for cell in row])
                blocks.append({"type": "table", "rows": rows, "page": sheet_index})
            page_count = len(workbook.worksheets)
        finally:
            workbook.close()

        processing_time_ms = round((perf_counter() - started) * 1000, 3)
        json_data = {
            "schema_version": SCHEMA_VERSION,
            "file_type": "xlsx",
            "page_count": page_count,
            "blocks": blocks,
            "meta": {
                "parser": "xlsx_parser",
                "processing_time_ms": processing_time_ms,
            },
        }
        return ParseResult(
            markdown="",
            json_data=json_data,
            page_count=page_count,
            processing_time_ms=processing_time_ms,
        )


ParserRegistry.register(XlsxParser())
