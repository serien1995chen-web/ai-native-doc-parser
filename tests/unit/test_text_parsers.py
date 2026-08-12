"""Unit tests for XLSX, HTML, text, and code parsers."""

from __future__ import annotations

import importlib
from pathlib import Path

import pytest
from openpyxl import Workbook

from app.parsers.implementations import code_parser as code_parser_module
from app.parsers.implementations import html_parser as html_parser_module
from app.parsers.implementations import text_parser as text_parser_module
from app.parsers.implementations import xlsx_parser as xlsx_parser_module
from app.parsers.implementations.code_parser import CodeParser
from app.parsers.implementations.html_parser import HtmlParser
from app.parsers.implementations.text_parser import TextParser
from app.parsers.implementations.xlsx_parser import XlsxParser
from app.parsers.registry import ParserRegistry

pytestmark = pytest.mark.unit


@pytest.fixture
def xlsx_path(tmp_path: Path) -> Path:
    """Create an XLSX with one populated sheet and one empty sheet."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Name", "Count", "Note"])
    sheet.append(["Alice", 3, None])
    sheet.append(["Bob", 7, ""])
    workbook.create_sheet("Empty")
    path = tmp_path / "sample.xlsx"
    workbook.save(path)
    workbook.close()
    return path


@pytest.fixture
def html_path(tmp_path: Path) -> Path:
    """Create an HTML file covering headings, lists, tables, images, and divs."""
    content = """<!doctype html>
<html><body>
<h1>Title</h1>
<p>Paragraph</p>
<ul><li>Item one</li><li>Item two</li></ul>
<table><tr><th>A</th><th>B</th></tr><tr><td>1</td><td>2</td></tr></table>
<div><img src="x.png" alt="x"></div>
<div><div>Nested text</div></div>
<div><p>Inside div</p></div>
</body></html>
"""
    path = tmp_path / "sample.html"
    path.write_text(content, encoding="utf-8")
    return path


@pytest.fixture
def txt_path(tmp_path: Path) -> Path:
    """Create a plain multi-line text file."""
    path = tmp_path / "sample.txt"
    path.write_text("line one\nline two\n", encoding="utf-8")
    return path


@pytest.fixture
def md_path(tmp_path: Path) -> Path:
    """Create a Markdown file with formatting that must not be rewritten."""
    content = "# Title\n\n**bold**\n\n```python\nprint(1)\n```\n"
    path = tmp_path / "sample.md"
    path.write_text(content, encoding="utf-8")
    return path


def _assert_common_schema(
    result,
    file_type: str,
    page_count: int,
    parser_name: str,
) -> None:
    assert result.markdown == ""
    assert result.page_count == page_count
    assert result.processing_time_ms is not None
    assert result.processing_time_ms >= 0

    json_data = result.json_data
    assert json_data["schema_version"] == "1.0"
    assert json_data["file_type"] == file_type
    assert json_data["page_count"] == page_count
    assert isinstance(json_data["blocks"], list)
    assert json_data["meta"]["parser"] == parser_name
    assert json_data["meta"]["processing_time_ms"] == result.processing_time_ms


def test_xlsx_parser_returns_blocks_schema(xlsx_path: Path) -> None:
    result = XlsxParser().parse(str(xlsx_path))
    _assert_common_schema(result, "xlsx", 2, "xlsx_parser")

    blocks = result.json_data["blocks"]
    headings = [block["text"] for block in blocks if block["type"] == "heading"]
    assert headings == ["Sheet Data", "Sheet Empty"]

    tables = [block for block in blocks if block["type"] == "table"]
    assert tables[0]["rows"] == [
        ["Name", "Count", "Note"],
        ["Alice", "3", ""],
        ["Bob", "7", ""],
    ]
    assert tables[1]["rows"] == []


def test_html_parser_returns_blocks_schema(html_path: Path) -> None:
    result = HtmlParser().parse(str(html_path))
    _assert_common_schema(result, "html", 1, "html_parser")

    blocks = result.json_data["blocks"]
    headings = [block for block in blocks if block["type"] == "heading"]
    assert headings[0]["text"] == "Title"
    assert headings[0]["level"] == 1

    paragraphs = [block["text"] for block in blocks if block["type"] == "paragraph"]
    assert "Paragraph" in paragraphs
    assert "Item one" in paragraphs
    assert "Item two" in paragraphs
    assert paragraphs.count("Nested text") == 1
    assert "Inside div" in paragraphs

    tables = [block for block in blocks if block["type"] == "table"]
    assert tables[0]["rows"] == [["A", "B"], ["1", "2"]]

    images = [block for block in blocks if block["type"] == "image"]
    assert len(images) == 1
    assert images[0]["bbox"] == []


def test_text_parser_preserves_content(txt_path: Path, md_path: Path) -> None:
    txt_result = TextParser().parse(str(txt_path))
    _assert_common_schema(txt_result, "txt", 1, "text_parser")
    assert txt_result.json_data["blocks"][0]["text"] == "line one\nline two\n"

    md_result = TextParser().parse(str(md_path))
    _assert_common_schema(md_result, "md", 1, "text_parser")
    assert (
        md_result.json_data["blocks"][0]["text"]
        == "# Title\n\n**bold**\n\n```python\nprint(1)\n```\n"
    )


def test_code_parser_detects_language(tmp_path: Path) -> None:
    cases = [
        ("sample.py", "print('hi')\n", "python"),
        ("sample.js", "console.log(1);\n", "javascript"),
        ("sample.json", '{"a": 1}\n', "json"),
    ]
    for filename, content, expected in cases:
        path = tmp_path / filename
        path.write_text(content, encoding="utf-8")
        result = CodeParser().parse(str(path))
        _assert_common_schema(result, "code", 1, "code_parser")
        block = result.json_data["blocks"][0]
        assert block["type"] == "code"
        assert block["language"] == expected
        assert block["text"] == content

    unknown = tmp_path / "notes.unknown"
    unknown.write_text("hello world\n", encoding="utf-8")
    unknown_result = CodeParser().parse(str(unknown))
    assert unknown_result.json_data["blocks"][0]["language"] == "text"


def test_text_parsers_registered() -> None:
    ParserRegistry._parsers.clear()
    importlib.reload(xlsx_parser_module)
    importlib.reload(html_parser_module)
    importlib.reload(text_parser_module)
    importlib.reload(code_parser_module)

    assert isinstance(ParserRegistry.get_parser("xls"), xlsx_parser_module.XlsxParser)
    assert isinstance(ParserRegistry.get_parser("xlsx"), xlsx_parser_module.XlsxParser)
    assert isinstance(ParserRegistry.get_parser("html"), html_parser_module.HtmlParser)
    assert isinstance(ParserRegistry.get_parser("htm"), html_parser_module.HtmlParser)
    assert isinstance(ParserRegistry.get_parser("txt"), text_parser_module.TextParser)
    assert isinstance(ParserRegistry.get_parser("md"), text_parser_module.TextParser)
    assert isinstance(ParserRegistry.get_parser("text"), text_parser_module.TextParser)
    assert isinstance(ParserRegistry.get_parser("code"), code_parser_module.CodeParser)
    assert isinstance(
        ParserRegistry.get_parser("python"), code_parser_module.CodeParser
    )
    assert isinstance(
        ParserRegistry.get_parser("javascript"), code_parser_module.CodeParser
    )


def test_text_parsers_raise_file_not_found(tmp_path: Path) -> None:
    missing = tmp_path / "missing-file"
    parsers = (XlsxParser(), HtmlParser(), TextParser(), CodeParser())
    for parser in parsers:
        with pytest.raises(FileNotFoundError):
            parser.parse(str(missing))


def test_empty_files_return_valid_schema(tmp_path: Path) -> None:
    txt_path = tmp_path / "empty.txt"
    txt_path.write_text("", encoding="utf-8")
    txt_result = TextParser().parse(str(txt_path))
    _assert_common_schema(txt_result, "txt", 1, "text_parser")
    assert txt_result.json_data["blocks"][0]["text"] == ""

    code_path = tmp_path / "empty.py"
    code_path.write_text("", encoding="utf-8")
    code_result = CodeParser().parse(str(code_path))
    _assert_common_schema(code_result, "code", 1, "code_parser")
    assert code_result.json_data["blocks"][0]["text"] == ""

    html_path = tmp_path / "empty.html"
    html_path.write_text("", encoding="utf-8")
    html_result = HtmlParser().parse(str(html_path))
    _assert_common_schema(html_result, "html", 1, "html_parser")
    assert html_result.json_data["blocks"] == []

    workbook = Workbook()
    workbook.active.title = "Empty"
    xlsx_path = tmp_path / "empty.xlsx"
    workbook.save(xlsx_path)
    workbook.close()
    xlsx_result = XlsxParser().parse(str(xlsx_path))
    _assert_common_schema(xlsx_result, "xlsx", 1, "xlsx_parser")
    assert xlsx_result.json_data["blocks"][0]["type"] == "heading"
    assert xlsx_result.json_data["blocks"][1]["type"] == "table"
    assert xlsx_result.json_data["blocks"][1]["rows"] == []
